import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .common import flatten_json
from .storage import persist_export_bytes


HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONO_FONT = Font(name="Consolas", size=10)


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _apply_wrap(cell):
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if cell.column == 1:
        cell.font = MONO_FONT


def _auto_width(ws):
    for column_cells in ws.columns:
        cells = list(column_cells)
        if not cells:
            continue
        column_letter = cells[0].column_letter
        max_length = 0
        for cell in cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _prepare_sheet(ws, headers):
    ws.append(headers)
    _style_header(ws[1])
    ws.freeze_panes = "A2"


def _finalize_sheet(ws):
    ws.auto_filter.ref = ws.dimensions


def _write_key_value_sheet(ws, items):
    _prepare_sheet(ws, ["Campo", "Valor"])
    for key, value in items:
        ws.append([key, "" if value is None else value])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            _apply_wrap(cell)
    _finalize_sheet(ws)


def _write_table_sheet(ws, headers, rows):
    _prepare_sheet(ws, headers)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            _apply_wrap(cell)
    _finalize_sheet(ws)


def _write_json_sheet(ws, data):
    _prepare_sheet(ws, ["Path", "Valor"])
    for path, value in flatten_json(data):
        ws.append([path, value])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            _apply_wrap(cell)
    _finalize_sheet(ws)


def _build_workbook(context):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    scenario = context.get("scenario") or {}
    report = context.get("dry_run") or {}
    answers = scenario.get("source_answers") or {}

    ws = workbook.create_sheet("Resumo")
    resumo_rows = [
        ("generated_at", context["metadata"]["generated_at"]),
        ("export_type", context["metadata"]["export_type"]),
        ("source_type", context["metadata"]["source_type"]),
        ("source_id", context["metadata"]["source_id"]),
        ("titulo", context.get("title", "")),
        ("resumo", context.get("summary", "")),
        ("campaign_id", answers.get("campaign_id", "")),
        ("campaign_name", answers.get("campaign_name", "")),
        ("customer_type", answers.get("customer_type", "")),
        ("document_type", answers.get("document_type", "")),
        ("event_type", answers.get("event_type", "")),
        ("objective", answers.get("objective", "")),
    ]
    if report:
        resumo_rows.extend(
            [
                ("dry_run_status", report.get("status", "")),
                ("dry_run_started_at", report.get("started_at", "")),
                ("dry_run_finished_at", report.get("finished_at", "")),
                ("dry_run_duration_ms", report.get("duration_ms", "")),
            ]
        )
    _write_key_value_sheet(ws, resumo_rows)

    if context.get("execution_steps"):
        ws = workbook.create_sheet("Execution Steps")
        rows = [
            [
                item.get("step", ""),
                item.get("action", ""),
                item.get("details", ""),
                item.get("expected_result", ""),
            ]
            for item in context["execution_steps"]
        ]
        _write_table_sheet(ws, ["Step", "Action", "Details", "Expected Result"], rows)

    if context.get("validation_steps"):
        ws = workbook.create_sheet("Validation Steps")
        rows = [
            [
                item.get("step", ""),
                item.get("validation", ""),
                item.get("details", ""),
                item.get("expected_result", ""),
            ]
            for item in context["validation_steps"]
        ]
        _write_table_sheet(ws, ["Step", "Validation", "Details", "Expected Result"], rows)

    ws = workbook.create_sheet("Payload")
    _write_json_sheet(ws, context.get("payload") or {})

    ws = workbook.create_sheet("Queries")
    query_rows = []
    for item in context.get("queries") or []:
        query_rows.append(
            [
                item.get("name", ""),
                item.get("kind", ""),
                item.get("purpose", ""),
                item.get("sql") or item.get("lookup") or "",
            ]
        )
    _write_table_sheet(ws, ["Name", "Kind", "Purpose", "Query"], query_rows)

    ws = workbook.create_sheet("Checkpoints")
    _write_table_sheet(ws, ["Checkpoint"], [[item] for item in context.get("checkpoints") or []])

    ws = workbook.create_sheet("Evidencias")
    _write_table_sheet(ws, ["Arquivo"], [[item] for item in context.get("evidence_files") or []])

    ws = workbook.create_sheet("Warnings")
    _write_table_sheet(ws, ["Warning"], [[item] for item in context.get("warnings") or []])

    if report:
        ws = workbook.create_sheet("Dry-run Results")
        summary = report.get("summary") or {}
        _write_key_value_sheet(
            ws,
            [
                ("report_id", report.get("id", "")),
                ("scenario_id", report.get("scenario_id", "")),
                ("status", report.get("status", "")),
                ("started_at", report.get("started_at", "")),
                ("finished_at", report.get("finished_at", "")),
                ("duration_ms", report.get("duration_ms", "")),
                ("total", summary.get("total", "")),
                ("passed", summary.get("passed", "")),
                ("failed", summary.get("failed", "")),
                ("skipped", summary.get("skipped", "")),
            ],
        )

        ws = workbook.create_sheet("Dry-run Logs")
        _write_table_sheet(ws, ["Log"], [[item] for item in report.get("logs") or []])

    for ws in workbook.worksheets:
        _auto_width(ws)

    return workbook


def export_xlsx_artifact(context):
    workbook = _build_workbook(context)
    buffer = BytesIO()
    workbook.save(buffer)
    return persist_export_bytes(
        buffer.getvalue(),
        context["source_type"],
        context["source_id"],
        context["metadata"]["export_type"],
        "xlsx",
    )
